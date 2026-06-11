"""
ColdSmart Report Generator Service
PDF, Excel, and CSV reports for compliance and business intelligence
"""
import io
import csv
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

# ReportLab for PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# OpenPyXL for Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

from app.models import (
    SensorReading, Alert, AlertSeverity, AuditLog,
    Device, Chamber, Report, ReportType, ReportFormat
)

# ─── Color palette (ColdSmart brand) ─────────────────────────────────────────
CS_BLUE = colors.HexColor("#1E3A5F")
CS_TEAL = colors.HexColor("#0EA5E9")
CS_GREEN = colors.HexColor("#22C55E")
CS_ORANGE = colors.HexColor("#F59E0B")
CS_RED = colors.HexColor("#EF4444")
CS_LIGHT = colors.HexColor("#F0F9FF")
CS_GRAY = colors.HexColor("#6B7280")


class ReportGenerator:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def generate_temperature_compliance_pdf(
        self,
        company_id: UUID,
        device_ids: List[UUID],
        date_from: datetime,
        date_to: datetime,
        title: str = "Temperature Compliance Report",
    ) -> bytes:
        """Generate a professional PDF temperature compliance report."""
        readings = await self._fetch_readings(device_ids, date_from, date_to)
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()
        story = []

        # Header
        story.append(self._make_header(title, date_from, date_to, styles))
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=2, color=CS_BLUE))
        story.append(Spacer(1, 0.5*cm))

        # Summary stats
        story.append(self._make_compliance_summary(readings, styles))
        story.append(Spacer(1, 0.5*cm))

        # Data table
        if readings:
            story.append(Paragraph("Temperature Log", styles["Heading2"]))
            story.append(Spacer(1, 0.3*cm))
            story.append(self._make_temperature_table(readings))

        story.append(Spacer(1, 1*cm))
        story.append(self._make_footer(styles))

        doc.build(story)
        return buffer.getvalue()

    async def generate_alert_report_pdf(
        self,
        company_id: UUID,
        device_ids: Optional[List[UUID]],
        date_from: datetime,
        date_to: datetime,
    ) -> bytes:
        alerts = await self._fetch_alerts(company_id, device_ids, date_from, date_to)
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story = []

        story.append(self._make_header("Alert Summary Report", date_from, date_to, styles))
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=2, color=CS_BLUE))
        story.append(Spacer(1, 0.5*cm))

        # Alert stats by severity
        total = len(alerts)
        emergency = sum(1 for a in alerts if a.severity == AlertSeverity.EMERGENCY)
        critical = sum(1 for a in alerts if a.severity == AlertSeverity.CRITICAL)
        warning = sum(1 for a in alerts if a.severity == AlertSeverity.WARNING)
        info = sum(1 for a in alerts if a.severity == AlertSeverity.INFO)

        summary_data = [
            ["Metric", "Count"],
            ["Total Alerts", str(total)],
            ["Emergency", str(emergency)],
            ["Critical", str(critical)],
            ["Warning", str(warning)],
            ["Info", str(info)],
        ]

        summary_table = Table(summary_data, colWidths=[8*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), CS_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#FEE2E2")),  # Emergency
            ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#FFEDD5")),  # Critical
            ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#FEF9C3")),  # Warning
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CS_LIGHT]),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.5*cm))

        # Alert details table
        if alerts:
            story.append(Paragraph("Alert Details", styles["Heading2"]))
            story.append(Spacer(1, 0.3*cm))

            headers = ["Time", "Severity", "Type", "Title", "Status"]
            table_data = [headers]
            for alert in alerts[:100]:  # Limit to 100 rows
                severity_str = alert.severity.value.upper()
                table_data.append([
                    alert.triggered_at.strftime("%Y-%m-%d %H:%M") if alert.triggered_at else "",
                    severity_str,
                    alert.alert_type.replace("_", " ").title(),
                    alert.title[:60] + "..." if len(alert.title) > 60 else alert.title,
                    alert.status.value.title(),
                ])

            col_widths = [3.5*cm, 2.5*cm, 4*cm, 6*cm, 2.5*cm]
            detail_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            detail_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), CS_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CS_LIGHT]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(detail_table)

        story.append(Spacer(1, 1*cm))
        story.append(self._make_footer(styles))
        doc.build(story)
        return buffer.getvalue()

    async def generate_excel_report(
        self,
        company_id: UUID,
        device_ids: List[UUID],
        date_from: datetime,
        date_to: datetime,
        report_type: ReportType,
    ) -> bytes:
        """Generate a comprehensive Excel workbook with charts."""
        readings = await self._fetch_readings(device_ids, date_from, date_to)

        wb = openpyxl.Workbook()

        # ── Summary Sheet ────────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._style_excel_header(ws_summary, "ColdSmart Compliance Report", date_from, date_to)

        # Stats
        ws_summary["A5"] = "Total Readings"
        ws_summary["B5"] = len(readings)
        ws_summary["A6"] = "Report Period"
        ws_summary["B6"] = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        ws_summary["A7"] = "Generated"
        ws_summary["B7"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        # ── Temperature Data Sheet ────────────────────────────────────────────
        ws_temp = wb.create_sheet("Temperature Data")
        headers = ["Timestamp", "Device", "Chamber", "Temperature (°C)", "Min Limit", "Max Limit", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_temp.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, reading in enumerate(readings, 2):
            ws_temp.cell(row=row_idx, column=1, value=reading.recorded_at.strftime("%Y-%m-%d %H:%M"))
            ws_temp.cell(row=row_idx, column=2, value=str(reading.device_id))
            ws_temp.cell(row=row_idx, column=3, value=str(reading.chamber_id))
            ws_temp.cell(row=row_idx, column=4, value=reading.temperature)
            ws_temp.cell(row=row_idx, column=5, value=None)  # min limit
            ws_temp.cell(row=row_idx, column=6, value=None)  # max limit

            # Conditional status
            temp = reading.temperature
            status = "OK" if temp else "N/A"
            ws_temp.cell(row=row_idx, column=7, value=status)

        # Auto-fit columns
        for col in ws_temp.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_temp.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

        # ── Chart ─────────────────────────────────────────────────────────────
        if len(readings) > 1:
            chart = LineChart()
            chart.title = "Temperature Over Time"
            chart.style = 10
            chart.y_axis.title = "Temperature (°C)"
            chart.x_axis.title = "Time"
            chart.width = 20
            chart.height = 10

            data_ref = Reference(ws_temp, min_col=4, min_row=1, max_row=min(len(readings)+1, 1000))
            chart.add_data(data_ref, titles_from_data=True)
            ws_temp.add_chart(chart, "I2")

        # ── Humidity Sheet ────────────────────────────────────────────────────
        ws_humid = wb.create_sheet("Humidity Data")
        humid_headers = ["Timestamp", "Chamber", "Humidity (%)", "Status"]
        for col, h in enumerate(humid_headers, 1):
            cell = ws_humid.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")

        for row_idx, reading in enumerate(readings, 2):
            ws_humid.cell(row=row_idx, column=1, value=reading.recorded_at.strftime("%Y-%m-%d %H:%M"))
            ws_humid.cell(row=row_idx, column=2, value=str(reading.chamber_id))
            ws_humid.cell(row=row_idx, column=3, value=reading.humidity)
            ws_humid.cell(row=row_idx, column=4, value="OK")

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def generate_csv_report(
        self,
        company_id: UUID,
        device_ids: List[UUID],
        date_from: datetime,
        date_to: datetime,
    ) -> bytes:
        """Generate a flat CSV with all sensor readings."""
        readings = await self._fetch_readings(device_ids, date_from, date_to)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "timestamp", "device_id", "chamber_id",
            "temperature_c", "humidity_pct", "co2_ppm",
            "o2_pct", "ethylene_ppm", "co_ppm", "methane_ppm",
            "health_score"
        ])

        for r in readings:
            writer.writerow([
                r.recorded_at.isoformat(),
                str(r.device_id),
                str(r.chamber_id),
                r.temperature,
                r.humidity,
                r.co2,
                r.o2,
                r.ethylene,
                r.carbon_monoxide,
                r.methane,
                r.health_score,
            ])

        return output.getvalue().encode("utf-8")

    # ─── Private helpers ──────────────────────────────────────────────────────

    async def _fetch_readings(self, device_ids: List[UUID], date_from: datetime, date_to: datetime):
        result = await self.db.execute(
            select(SensorReading)
            .where(
                SensorReading.device_id.in_(device_ids),
                SensorReading.recorded_at >= date_from,
                SensorReading.recorded_at <= date_to,
            )
            .order_by(SensorReading.recorded_at.asc())
            .limit(10000)
        )
        return result.scalars().all()

    async def _fetch_alerts(self, company_id: UUID, device_ids: Optional[List[UUID]], date_from: datetime, date_to: datetime):
        query = select(Alert).where(
            Alert.company_id == company_id,
            Alert.triggered_at >= date_from,
            Alert.triggered_at <= date_to,
        )
        if device_ids:
            query = query.where(Alert.device_id.in_(device_ids))
        result = await self.db.execute(query.order_by(Alert.triggered_at.desc()))
        return result.scalars().all()

    def _make_header(self, title: str, date_from: datetime, date_to: datetime, styles) -> Table:
        header_data = [
            [Paragraph(f"<b>{title}</b>", ParagraphStyle("H", fontSize=16, textColor=CS_BLUE, alignment=TA_LEFT)),
             Paragraph(
                 f"Period: {date_from.strftime('%d %b %Y')} – {date_to.strftime('%d %b %Y')}<br/>"
                 f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}",
                 ParagraphStyle("sub", fontSize=9, textColor=CS_GRAY, alignment=TA_RIGHT)
             )],
        ]
        t = Table(header_data, colWidths=["60%", "40%"])
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        return t

    def _make_compliance_summary(self, readings, styles) -> Table:
        temps = [r.temperature for r in readings if r.temperature is not None]
        if not temps:
            return Paragraph("No data available for the selected period.", styles["Normal"])

        avg_temp = sum(temps) / len(temps)
        min_temp = min(temps)
        max_temp = max(temps)

        data = [
            ["Metric", "Value"],
            ["Total Readings", str(len(readings))],
            ["Average Temperature", f"{avg_temp:.2f}°C"],
            ["Minimum Temperature", f"{min_temp:.2f}°C"],
            ["Maximum Temperature", f"{max_temp:.2f}°C"],
        ]

        t = Table(data, colWidths=[8*cm, 5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), CS_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CS_LIGHT]),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ]))
        return t

    def _make_temperature_table(self, readings) -> Table:
        headers = ["Timestamp", "Chamber", "Temp (°C)", "Humidity (%)", "CO₂ (ppm)", "Status"]
        data = [headers]
        for r in readings[:500]:
            data.append([
                r.recorded_at.strftime("%Y-%m-%d %H:%M") if r.recorded_at else "",
                str(r.chamber_id)[:8] + "...",
                f"{r.temperature:.1f}" if r.temperature else "N/A",
                f"{r.humidity:.1f}" if r.humidity else "N/A",
                f"{r.co2:.0f}" if r.co2 else "N/A",
                "OK",
            ])

        col_widths = [4*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), CS_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CS_LIGHT]),
        ]))
        return t

    def _make_footer(self, styles) -> Paragraph:
        return Paragraph(
            "This report is generated automatically by ColdSmart. "
            "All data is logged in real-time and audited. "
            "For queries, contact support@coldsmart.io",
            ParagraphStyle("footer", fontSize=7, textColor=CS_GRAY, alignment=TA_CENTER),
        )

    def _style_excel_header(self, ws, title: str, date_from: datetime, date_to: datetime):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
        ws["A2"] = f"Period: {date_from.strftime('%d %b %Y')} – {date_to.strftime('%d %b %Y')}"
        ws["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}"
        ws["A3"].font = Font(color="6B7280", size=9)
